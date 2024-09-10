from Script.Config import Config_Setup
import os
import shutil
import yaml
import random
import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def clean_dir(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)  # Remove file or link
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)  # Remove directory and its contents
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")


def get_first_file_name():
    files_list = []
    for filename in os.listdir(downloads_path):
        file_path = os.path.join(downloads_path, filename)
        print(file_path)
        files_list.append(filename)
        print(filename)
    return files_list[0]


def get_files_list(path):
    files_list = []
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        files_list.append(filename)
    return files_list


def take_backup(source_file_path, destination_dir):
    today_date = datetime.datetime.today().strftime('%d-%m-%Y')
    file_name, file_extension = os.path.splitext(os.path.basename(source_file_path))
    new_file_name = f"{file_name}_{today_date}{file_extension}"
    new_file_path = os.path.join(destination_dir, new_file_name)
    shutil.copy2(source_file_path, new_file_path)
    return new_file_path


def get_file_fullName(partial_file_name):
    sheets = get_files_list(Config_Setup.input_sheets_dir)
    for sheet in sheets:
        if partial_file_name in sheet:
            return sheet


def is_valid_excel_date(serial_date):
    # Excel uses 1-based serial dates from 1/1/1900 (serial number 1) to 12/31/9999
    return 1 <= serial_date <= 2958465  # 2958465 corresponds to 12/31/9999


def clear_filters_and_unhide_rows(file_path, sheet_name=None):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(file_path)

    # Select the active sheet or the specific sheet if provided
    sheet = wb.active if sheet_name is None else wb[sheet_name]

    # Clear any filters
    sheet.auto_filter = None

    # Unhide all rows
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].hidden = False

    # Save the workbook after modification
    wb.save(file_path)
    print(f"Filters cleared and all rows are unhidden in sheet: {sheet.title}")
