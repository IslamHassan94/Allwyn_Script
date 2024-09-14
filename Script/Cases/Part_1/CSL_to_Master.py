from datetime import datetime
from Script.Config import Config_Setup
from Script.Utils import FilesUtil
from Script.Data import ExportOrders
from Script.Utils import ProgressAnimation, DateUtil
from Script.Config import Logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
import logging
import pandas as pd
import xlwings as xw
import time
import threading
import sys

#################### Logging #############################
Logger.init_Logger()
logger = logging.getLogger("Allwyn Script")
logging.getLogger("imported_module").setLevel(logging.WARNING)
##################################################################
logger.debug("Getting Sheets names...")
vodafone_provide_path = Config_Setup.input_sheets_dir + FilesUtil.get_file_fullName_by_keyword_in_name(
    Config_Setup.vodafone_provide, 'Base')
site_status_report_path = Config_Setup.input_sheets_dir + FilesUtil.get_file_fullName(Config_Setup.site_status_report)
allwyn_fault_tracking_path = Config_Setup.input_sheets_dir + FilesUtil.get_file_fullName(
    Config_Setup.allwyn_fault_tracking)
orders = []
filtered_orders_without_true = []
true_orders = []


def handle_Csl_to_master():
    logger.debug("Exporting orders from Status Report Sheet...")
    orders = ExportOrders.export_orders_from_status_report()
    true_orders = ExportOrders.filter_True_orders(orders)
    filtered_orders_without_true = ExportOrders.filter_orders_without_true(orders, true_orders)
    print(f'orders count: {len(orders)}')
    print(f'true orders count: {len(true_orders)}')
    print(f'filtered orders count: {len(filtered_orders_without_true)}')
    df_master = pd.read_excel(vodafone_provide_path, sheet_name='Provide Update', skiprows=4)
    stop_event = threading.Event()
    progress_thread = threading.Thread(target=ProgressAnimation.rolling_progress_bar, args=(stop_event,))
    progress_thread.start()
    write_orders_to_master_sheet(filtered_orders_without_true, true_orders, df_master, vodafone_provide_path,
                                 'Provide Update')
    stop_event.set()
    progress_thread.join()
    logger.debug("\nData Transfer Completed.")


def write_orders_to_master_sheet(filtered_orders_not_true, true_orders, df_master, file_path, sheet_name):
    # Use xlwings to avoid Excel corruption
    app = xw.App(visible=False)
    book = app.books.open(file_path)
    sheet = book.sheets[sheet_name]

    # Create a dictionary to map retailer IDs to their row numbers in Excel
    retailer_row_map = {}
    for row in range(6, sheet.api.UsedRange.Rows.Count + 1):  # Assuming data starts from row 6 in Excel
        retailer_id_cell = sheet.range(f'A{row}').value  # Replace 'A' with the column that contains 'Retailer ID'
        if retailer_id_cell:
            retailer_row_map[retailer_id_cell] = row

    # Handle invalid date cells across all rows
    handle_invalid_dates(sheet)

    # Loop through each order in true_orders and update corresponding row in the Excel sheet
    for order in true_orders:
        if order.retailer_id in retailer_row_map:
            row_num = retailer_row_map[order.retailer_id]

            # Update the corresponding cells in the Excel sheet
            sheet.range(f'X{row_num}').value = order.date_required  # Replace 'X' with the appropriate column
            sheet.range(
                f'Y{row_num}').value = order.install_date  # Replace 'Y' with the column for 'Forecasted OLO install Date'
            converted_first_poll_date = datetime.strptime(order.first_poll_date, "%d/%m/%Y")
            sheet.range(f'Z{row_num}').value = converted_first_poll_date
            # # Handle first_poll_date to check if it's a datetime object or a string
            # if isinstance(order.first_poll_date, datetime.datetime):
            #     sheet.range(f'Z{row_num}').value = order.first_poll_date.strftime('%d/%m/%y')
            # else:
            #     sheet.range(f'Z{row_num}').value = order.first_poll_date  # Directly assign if it's already a string

            print(sheet.range(f'Z{row_num}').value)

            # Safely handle None values for DateUtil.getTodaysDate()
            today_date = DateUtil.getTodaysDate()

            # Write order body and message only if they are not NaN, empty, or whitespace
            order_body = order.body or ''
            order_message = order.message or ''

            if pd.notna(order_body) and order_body.strip() or pd.notna(order_message) and order_message.strip():
                # Write today's date in column 'I' only if body or message is not empty
                sheet.range(f'I{row_num}').value = today_date

                # Write order body and message in columns 'M' and 'L' respectively if not empty
                if pd.notna(order_body) and order_body.strip():
                    sheet.range(f'M{row_num}').value = order_body
                if pd.notna(order_message) and order_message.strip():
                    sheet.range(f'L{row_num}').value = order_message

                # Update concatenated messages in columns 'K' and 'J' based on body and message content
                if pd.notna(order_body) and order_body.strip() and not pd.notna(order_message):
                    concatenated_message = f"{today_date}: {order_body}"
                    sheet.range(f'K{row_num}').value = concatenated_message
                elif pd.notna(order_message) and order_message.strip() and not pd.notna(order_body):
                    concatenated_message = f"{today_date}: {order_message}"
                    sheet.range(f'K{row_num}').value = concatenated_message
                elif pd.notna(order_message) and order_message.strip() and pd.notna(order_body) and order_body.strip():
                    concatenated_message = f"{today_date}: {order_body} {order_message}"
                    sheet.range(f'K{row_num}').value = concatenated_message

                # Copy the value from column 'K' to column 'J'
                sheet.range(f'J{row_num}').value = sheet.range(f'K{row_num}').value
                sheet.range(f'AA{row_num}').value = 'TRUE'
                sheet.range(f'U{row_num}').value = '2.Commissioning'

    # New loop to handle filtered_orders_not_true
    for order in filtered_orders_not_true:
        if order.retailer_id in retailer_row_map:
            row_num = retailer_row_map[order.retailer_id]
            # Only update columns X and Y
            sheet.range(f'X{row_num}').value = order.date_required  # Replace 'X' with the appropriate column
            sheet.range(
                f'Y{row_num}').value = order.install_date  # Replace 'Y' with the column for 'Forecasted OLO install Date'

    # Save and close the workbook without Excel recovery prompts
    book.save()
    book.close()
    app.quit()


def handle_invalid_dates(sheet):
    # Iterate through all rows and check for invalid date serial values in column 'T'
    for row in range(6, sheet.api.UsedRange.Rows.Count + 1):  # Adjust start row as needed
        cell_value = sheet.range(f'T{row}').value  # Adjust column 'T' if necessary
        if isinstance(cell_value, int) and cell_value > 2958465:  # Max valid Excel date serial
            print(f"Invalid date value in cell T{row}: {cell_value}")
            # Handle the invalid date by setting it to None
            sheet.range(f'T{row}').value = None


def generate_final_vodafone_provide_sheet(path):
    # Define the required columns
    required_columns = [
        'Retailer ID', 'SR No.', 'Order batch date', 'Allwyn Site Type (ie Type 1 or 2)',
        'SOGEA / FTTP', 'Store Name', 'City', 'Postcode', 'Updates / Comments',
        'Access Service Id (VF Access Service Id)', 'Connection ID (CSL Service)', 'Site Status',
        'Appointment Slot - AM (9am-1pm) / PM (1pm - 5pm)', 'Site Survey Date', 'Initial OLO requested date',
        'Forecasted OLO install Date', 'Actual completion date OLO First Poll Date',
        'OLO Service Activated', 'Line test (Fault)', 'Scheduled Router Install Date',
        'Completed Router Install Date & Time', 'CSL Router - S/N'
    ]

    # Read the Excel sheet
    try:
        df = pd.read_excel(path, sheet_name='Provide Update', skiprows=4)
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None

    # Replace newline characters in column names for easier handling
    df.columns = df.columns.str.replace('\n', ' ').str.strip()

    # Check if the required columns are present in the DataFrame
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise KeyError(f"The following columns are missing from the DataFrame: {missing_columns}")

    # Filter the DataFrame to include only the required columns
    filtered_df = df.loc[:, required_columns]

    # Define columns that potentially contain dates
    date_columns = [
        'Site Survey Date', 'Initial OLO requested date', 'Forecasted OLO install Date',
        'Actual completion date OLO First Poll Date',
        'Scheduled Router Install Date', 'Completed Router Install Date & Time'
    ]

    # Convert dates to datetime and format them
    for col in date_columns:
        if col in filtered_df.columns:
            filtered_df[col] = pd.to_datetime(filtered_df[col], errors='coerce', dayfirst=True)
            filtered_df[col] = filtered_df[col].dt.strftime('%d/%m/%Y')

    # Save the filtered DataFrame to a new Excel file
    output_file_path = Config_Setup.output_folder + f'Vodafone_Provide_Update_{DateUtil.getTodaysDateInSerialFormat()}.xlsx'
    try:
        filtered_df.to_excel(output_file_path, index=False)
    except Exception as e:
        print(f"Error saving the file: {e}")


def add_group_by_month_filter():
    output_file_path = Config_Setup.output_folder + f'Vodafone_Provide_Update_{DateUtil.getTodaysDateInSerialFormat()}.xlsx'

    # Read the Excel file into a DataFrame
    df = pd.read_excel(output_file_path)

    # Convert the 'Actual completion date OLO First Poll Date' to datetime and format as 'dd/mm/yyyy'
    df['Actual completion date OLO First Poll Date'] = pd.to_datetime(df['Actual completion date OLO First Poll Date'],
                                                                      dayfirst=True, errors='coerce')

    # Ensure the date is formatted as 'dd/mm/yyyy'
    df['Actual completion date OLO First Poll Date'] = df['Actual completion date OLO First Poll Date'].dt.strftime(
        '%d/%m/%Y')

    # Save the DataFrame back to the same Excel file
    with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False)

    # Load the workbook and worksheet using openpyxl
    wb = load_workbook(output_file_path)
    ws = wb.active

    # Apply filter on the column "Actual completion date OLO First Poll Date"
    ws.auto_filter.ref = ws.dimensions  # Auto filter for the entire sheet
    ws.auto_filter.add_filter_column(0, [])  # Apply filter to the first column (change index if necessary)

    # Define a date style for the "dd/mm/yyyy" format
    date_style = NamedStyle(name="dd_mm_yyyy", number_format="DD/MM/YYYY")

    # Apply the date style to the relevant column
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):  # Assuming date is in column 1
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.style = date_style

    # Save the updated workbook with the filter and date formatting
    wb.save(output_file_path)
